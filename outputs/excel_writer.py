"""
excel_writer.py - Export approved actions + the dashboard to Excel.

Writes outputs/acciones_aprobadas.xlsx with:
  - sheet "acciones_aprobadas": one row per APPROVED or EDITED action, including
    the final message (mensaje_final) and the client's contact channel so the
    broker can act on it. Pending and rejected actions are NOT exported.
  - sheet "dashboard": the deterministic portfolio metrics + the insight narrative.
  - sheets "mix_aseguradora" / "mix_ramo": the portfolio mix tables.

Every value written here was computed by cartera_core or decided by the broker.
This module formats; it never computes a domain number.
"""

import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
NEXO = os.path.dirname(HERE)
if NEXO not in sys.path:
    sys.path.insert(0, NEXO)

import paths
import review

ACCION_COLS = [
    "tipo", "severidad", "confianza_%", "cliente_id", "cliente_nombre",
    "email", "telefono", "poliza", "detalle", "mensaje_final", "estado",
    "decidido_por", "decision_note", "ts_creada", "ts_decidida",
]


def _accion_rows(ctx):
    rows = []
    for d in review.approved_for_export(ctx):
        rows.append({
            "tipo": d["tipo"], "severidad": d["severidad"],
            "confianza_%": round(d["confianza"] * 100),
            "cliente_id": d["cliente_id"], "cliente_nombre": d["cliente_nombre"],
            "email": d.get("email"), "telefono": d.get("telefono"),
            "poliza": d.get("poliza"), "detalle": d["detalle"],
            "mensaje_final": d.get("mensaje_final") or d.get("mensaje_propuesto"),
            "estado": d["estado"], "decidido_por": d.get("decided_by"),
            "decision_note": d.get("decision_note", ""),
            "ts_creada": d.get("ts_creada"), "ts_decidida": d.get("ts_decidida"),
        })
    # Stable, useful order: by severity then confidence (mirrors the inbox).
    rows.sort(key=lambda r: (review.SEVERITY_ORDER.get(r["severidad"], 9), -r["confianza_%"]))
    return rows


def _dashboard_rows(ctx):
    m = ctx.state.get("metrics") or ctx.get("analisis_cartera_agent", "metrics", {}) or {}
    s = review.summary(ctx)
    narrative = ctx.get("analisis_cartera_agent", "narrative", "")
    venc_dias = m.get("vencimientos_dias", 30)
    metric_rows = [
        ("Total de pólizas", m.get("total_polizas")),
        ("Pólizas activas", m.get("polizas_activas")),
        ("Pólizas vencidas", m.get("polizas_vencidas")),
        ("Pólizas canceladas", m.get("polizas_canceladas")),
        ("Total de clientes", m.get("total_clientes")),
        ("Clientes activos", m.get("clientes_activos")),
        ("Clientes inactivos", m.get("clientes_inactivos")),
        ("Retención de clientes (%)", m.get("retencion_pct")),
        ("Prima mensual activa (ARS)", m.get("prima_mensual_total")),
        ("Comisión mensual estimada (ARS)", m.get("comision_mensual_total")),
        ("Pólizas en mora", m.get("polizas_en_mora")),
        ("Mora sobre pólizas activas (%)", m.get("pct_en_mora_polizas")),
        ("Mora sobre prima (%)", m.get("pct_en_mora_prima")),
        (f"Vencimientos próximos ({venc_dias} días)", m.get("vencimientos_proximos")),
        ("", ""),
        ("Acciones propuestas (total)", s["total"]),
        ("Aprobadas / editadas (exportadas)", s["exportables"]),
        ("Rechazadas", s["by_estado"].get(review.RECHAZADA, 0)),
        ("Pendientes", s["pendientes"]),
        ("", ""),
        ("Insight del panel", narrative),
    ]
    return [{"metrica": k, "valor": v} for k, v in metric_rows]


def _mix_rows(ctx, key):
    m = ctx.state.get("metrics") or ctx.get("analisis_cartera_agent", "metrics", {}) or {}
    mix = m.get(key, {})
    label = "aseguradora" if "aseg" in key else "ramo"
    rows = [{label: name, "polizas": v["count"], "prima_mensual": v["prima_mensual"]}
            for name, v in mix.items()]
    rows.sort(key=lambda r: -r["prima_mensual"])
    return rows


def _autosize(ws, max_width=80):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, max_width)


def export(ctx, path=None):
    """Write the Excel and return its path. Only approved/edited rows are exported."""
    import pandas as pd
    from openpyxl.styles import Font

    path = path or paths.APPROVED_XLSX
    paths.ensure_dirs()

    acciones = pd.DataFrame(_accion_rows(ctx), columns=ACCION_COLS)
    dashboard = pd.DataFrame(_dashboard_rows(ctx), columns=["metrica", "valor"])
    mix_aseg = pd.DataFrame(_mix_rows(ctx, "mix_por_aseguradora"),
                            columns=["aseguradora", "polizas", "prima_mensual"])
    mix_ramo = pd.DataFrame(_mix_rows(ctx, "mix_por_ramo"),
                            columns=["ramo", "polizas", "prima_mensual"])

    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        acciones.to_excel(xw, sheet_name="acciones_aprobadas", index=False)
        dashboard.to_excel(xw, sheet_name="dashboard", index=False)
        mix_aseg.to_excel(xw, sheet_name="mix_aseguradora", index=False)
        mix_ramo.to_excel(xw, sheet_name="mix_ramo", index=False)
        for name, ws in xw.sheets.items():
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"
            _autosize(ws)
            # wrap the long message / narrative columns
            from openpyxl.styles import Alignment
            for col in ws.columns:
                header = col[0].value
                if header in ("mensaje_final", "valor"):
                    ws.column_dimensions[col[0].column_letter].width = 70
                    for c in col[1:]:
                        c.alignment = Alignment(wrap_text=True, vertical="top")
    return path


if __name__ == "__main__":
    # Standalone: run the orchestrator (auto) and export.
    import nexo_orchestrator as orch
    os.environ.setdefault("NEXO_AUTO_APPROVE", "1")
    ctx = orch.run()
    print("\nExportado:", paths.APPROVED_XLSX)
