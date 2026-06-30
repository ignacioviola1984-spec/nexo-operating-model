"""Fail-closed Excel ingestion (§6) - as rigorous as a database load.

Pipeline: read all sheets -> validate (collecting EVERY error, not stopping at the
first) -> if any blocking error, reject the WHOLE file with a readable Spanish
report and change nothing (the prior snapshot stays active) -> on success,
materialize an immutable dated snapshot and archive the prior one.

A bad or incomplete upload is rejected whole, never ingested partially.
"""

from __future__ import annotations

import hashlib
import json
import math
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError

from nexo_os.data.repository import NexoRepository
from nexo_os.data.schema.models import (
    MODEL_BY_TABLE,
    OPERATIONAL_TABLES,
    OPTIONAL_TABLES,
    DataSnapshot,
    EstadoSnapshot,
    _Row,
)

# Natural primary key per operational table.
PK_BY_TABLE: dict[str, str] = {
    "clientes": "cliente_id",
    "polizas": "poliza_id",
    "cuotas": "cuota_id",
    "comisiones": "comision_id",
    "leads": "lead_id",
    "cotizaciones": "cotizacion_id",
    "aseguradoras": "aseguradora_id",
    "productores": "productor_id",
    "siniestros": "siniestro_id",
}

# Foreign keys: (table, fk_field, ref_table, required). required=False -> only
# checked when the value is present (nullable FK).
FOREIGN_KEYS: tuple[tuple[str, str, str, bool], ...] = (
    ("clientes", "productor_id", "productores", True),
    ("polizas", "cliente_id", "clientes", True),
    ("polizas", "aseguradora_id", "aseguradoras", True),
    ("polizas", "productor_id", "productores", True),
    ("polizas", "poliza_origen_id", "polizas", False),
    ("cuotas", "poliza_id", "polizas", True),
    ("comisiones", "poliza_id", "polizas", True),
    ("comisiones", "aseguradora_id", "aseguradoras", True),
    ("leads", "productor_id", "productores", True),
    ("leads", "cliente_id", "clientes", False),
    ("cotizaciones", "lead_id", "leads", True),
    ("cotizaciones", "aseguradora_id", "aseguradoras", True),
    ("cotizaciones", "poliza_id", "polizas", False),
    ("siniestros", "poliza_id", "polizas", True),
)

_PYDANTIC_ES: dict[str, str] = {
    "missing": "Campo requerido faltante o vacio.",
    "enum": "Valor fuera del dominio permitido.",
    "decimal_parsing": "No es un numero (decimal) valido.",
    "decimal_type": "No es un numero (decimal) valido.",
    "int_parsing": "No es un numero entero valido.",
    "int_type": "No es un numero entero valido.",
    "float_parsing": "No es un numero valido.",
    "date_parsing": "Fecha invalida (use AAAA-MM-DD).",
    "date_from_datetime_parsing": "Fecha invalida (use AAAA-MM-DD).",
    "date_type": "Fecha invalida (use AAAA-MM-DD).",
    "bool_parsing": "Valor booleano invalido (use TRUE/FALSE).",
    "string_type": "Texto invalido.",
    "extra_forbidden": "Columna no reconocida para esta hoja.",
}


@dataclass(frozen=True)
class ValidationIssue:
    sheet: str
    codigo: str
    mensaje: str
    fila: int | None = None  # 1-based Excel row (header is row 1)
    columna: str | None = None


@dataclass
class ValidationReport:
    errores: list[ValidationIssue] = field(default_factory=list)
    advertencias: list[ValidationIssue] = field(default_factory=list)
    row_counts: dict[str, int] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        """Fail-closed: any blocking error rejects the whole file."""
        return not self.errores

    def render_es(self) -> str:
        """Human-readable Spanish report (errors by sheet, row, column)."""
        lines: list[str] = []
        if self.ok:
            lines.append("VALIDACION OK - el archivo es apto para ingesta.")
        else:
            lines.append(
                f"ARCHIVO RECHAZADO - {len(self.errores)} error(es). "
                "No se ingirio nada; el snapshot anterior sigue activo."
            )
        if self.row_counts:
            counts = ", ".join(f"{t}={n}" for t, n in self.row_counts.items())
            lines.append(f"Filas leidas: {counts}")
        if self.errores:
            lines.append("")
            lines.append("ERRORES (bloqueantes):")
            for iss in self.errores:
                loc = []
                if iss.fila is not None:
                    loc.append(f"fila {iss.fila}")
                if iss.columna:
                    loc.append(f"col '{iss.columna}'")
                loc_s = f" ({', '.join(loc)})" if loc else ""
                lines.append(f"  - [{iss.sheet}]{loc_s}: {iss.mensaje}")
        if self.advertencias:
            lines.append("")
            lines.append("ADVERTENCIAS (no bloqueantes):")
            for iss in self.advertencias:
                lines.append(f"  - [{iss.sheet}]: {iss.mensaje}")
        return "\n".join(lines)


@dataclass
class IngestResult:
    report: ValidationReport
    snapshot: DataSnapshot | None = None

    @property
    def ok(self) -> bool:
        return self.report.ok and self.snapshot is not None


# --------------------------------------------------------------------------- #
# Reading
# --------------------------------------------------------------------------- #
def _clean(value: object) -> object:
    """Normalize a raw cell value for typed validation."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return value
    s = str(value).strip()
    return s or None


def read_workbook(path: Path) -> dict[str, list[dict]]:
    """Read every sheet into {sheet: [row_dict, ...]} with cleaned cell values.

    Each row_dict carries a private '__row__' key with the 1-based Excel row.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets: dict[str, list[dict]] = {}
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            sheets[ws.title] = []
            continue
        headers = [str(h).strip() if h is not None else "" for h in header]
        out: list[dict] = []
        for excel_row, raw in enumerate(rows_iter, start=2):
            if raw is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
                continue  # skip fully-empty rows
            row = {headers[i]: _clean(v) for i, v in enumerate(raw) if i < len(headers)}
            row["__row__"] = excel_row
            out.append(row)
        sheets[ws.title] = out
    wb.close()
    return sheets


# --------------------------------------------------------------------------- #
# Validation
# --------------------------------------------------------------------------- #
def _decimal_fields(model_cls: type[_Row]) -> list[str]:
    fields = []
    for name, info in model_cls.model_fields.items():
        ann = info.annotation
        if ann is Decimal or (hasattr(ann, "__args__") and Decimal in getattr(ann, "__args__", ())):
            fields.append(name)
    return fields


def validate_workbook(
    sheets: dict[str, list[dict]],
) -> tuple[ValidationReport, dict[str, list[_Row]]]:
    """Validate all sheets, collecting every error. Returns (report, typed_data)."""
    report = ValidationReport()
    typed: dict[str, list[_Row]] = {t: [] for t in OPERATIONAL_TABLES}

    # 1) required sheets present
    for table in OPERATIONAL_TABLES:
        if table not in sheets:
            if table in OPTIONAL_TABLES:
                report.advertencias.append(
                    ValidationIssue(table, "hoja_opcional_ausente", "Hoja opcional ausente.")
                )
            else:
                report.errores.append(
                    ValidationIssue(table, "hoja_faltante", "Falta una hoja requerida.")
                )

    # 2) headers + 3) per-row typed construction
    for table in OPERATIONAL_TABLES:
        if table not in sheets:
            continue
        model_cls = MODEL_BY_TABLE[table]
        expected = list(model_cls.model_fields.keys())
        rows = sheets[table]
        report.row_counts[table] = len(rows)

        present_headers = set()
        if rows:
            present_headers = {k for k in rows[0] if k != "__row__"}
        elif sheets[table] == [] and table in sheets:
            # Empty sheet: we cannot read headers from data; skip header check.
            present_headers = set(expected)

        missing_cols = [c for c in expected if c not in present_headers] if rows else []
        for col in missing_cols:
            report.errores.append(
                ValidationIssue(
                    table, "columna_faltante", f"Falta la columna '{col}'.", columna=col
                )
            )
        for col in present_headers - set(expected):
            report.advertencias.append(
                ValidationIssue(
                    table, "columna_extra", f"Columna no reconocida '{col}'.", columna=col
                )
            )

        if missing_cols:
            continue  # cannot reliably type rows without all columns

        dec_fields = _decimal_fields(model_cls)
        for row in rows:
            excel_row = row.get("__row__")
            payload = {k: v for k, v in row.items() if k != "__row__"}
            try:
                obj = model_cls(**payload)
            except ValidationError as exc:
                for err in exc.errors():
                    col = ".".join(str(p) for p in err["loc"]) if err["loc"] else None
                    msg = _PYDANTIC_ES.get(err["type"], err.get("msg", "Valor invalido."))
                    report.errores.append(
                        ValidationIssue(table, err["type"], msg, fila=excel_row, columna=col)
                    )
                continue
            # 4a) amount sign checks (models stay permissive; signs checked here)
            for fld in dec_fields:
                val = getattr(obj, fld)
                if isinstance(val, Decimal) and val < 0:
                    report.errores.append(
                        ValidationIssue(
                            table,
                            "monto_negativo",
                            f"'{fld}' no puede ser negativo.",
                            fila=excel_row,
                            columna=fld,
                        )
                    )
            typed[table].append(obj)

    # 4b) cross-field date sanity (polizas vigencia)
    for obj in typed.get("polizas", []):
        if obj.fecha_fin_vigencia < obj.fecha_inicio_vigencia:
            report.errores.append(
                ValidationIssue(
                    "polizas",
                    "fechas_incoherentes",
                    f"poliza {obj.poliza_id}: fecha_fin_vigencia < fecha_inicio_vigencia.",
                    columna="fecha_fin_vigencia",
                )
            )

    # 4c) PK uniqueness
    pk_sets: dict[str, set[str]] = {}
    for table in OPERATIONAL_TABLES:
        pk = PK_BY_TABLE[table]
        seen: set[str] = set()
        dupes: set[str] = set()
        for obj in typed.get(table, []):
            key = getattr(obj, pk)
            if key in seen:
                dupes.add(key)
            seen.add(key)
        pk_sets[table] = seen
        for d in sorted(dupes):
            report.errores.append(
                ValidationIssue(table, "pk_duplicada", f"{pk} duplicado: '{d}'.", columna=pk)
            )

    # 4d) referential integrity
    for table, fk_field, ref_table, required in FOREIGN_KEYS:
        ref_pks = pk_sets.get(ref_table, set())
        for obj in typed.get(table, []):
            val = getattr(obj, fk_field, None)
            if val is None:
                if required:
                    report.errores.append(
                        ValidationIssue(
                            table, "fk_nula", f"{fk_field} requerido.", columna=fk_field
                        )
                    )
                continue
            if val not in ref_pks:
                report.errores.append(
                    ValidationIssue(
                        table,
                        "fk_rota",
                        f"{fk_field}='{val}' no existe en {ref_table}.",
                        columna=fk_field,
                    )
                )

    return report, typed


# --------------------------------------------------------------------------- #
# Ingest
# --------------------------------------------------------------------------- #
def _file_hash(path: Path) -> str:
    h = hashlib.sha256()
    h.update(Path(path).read_bytes())
    return h.hexdigest()


def ingest(
    path: Path,
    *,
    cargado_por: str,
    repo: NexoRepository,
    snapshot_fecha: date,
    now: datetime,
    snapshot_id: str | None = None,
) -> IngestResult:
    """Validate `path` and, only if fully valid, materialize a new active snapshot.

    Fail-closed: on any blocking error nothing is written and the prior snapshot
    stays active. `repo` must expose `materialize_snapshot` (SnapshotRepository).
    """
    path = Path(path)
    sheets = read_workbook(path)
    report, typed = validate_workbook(sheets)
    if not report.ok:
        return IngestResult(report=report, snapshot=None)

    archivo_hash = _file_hash(path)
    sid = snapshot_id or f"snap-{snapshot_fecha:%Y%m%d}-{archivo_hash[:8]}"
    snapshot = DataSnapshot(
        snapshot_id=sid,
        snapshot_fecha=snapshot_fecha,
        archivo_nombre=path.name,
        archivo_hash=archivo_hash,
        cargado_por=cargado_por,
        cargado_en=now,
        row_counts_json=json.dumps(report.row_counts),
        estado=EstadoSnapshot.activo,
    )
    repo.materialize_snapshot(snapshot, typed)  # type: ignore[attr-defined]
    return IngestResult(report=report, snapshot=snapshot)
