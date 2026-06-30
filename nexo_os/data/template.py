"""Emit the canonical Excel template the broker fills (`make template`).

One sheet per operational table, headers EXACTLY matching the schema column
names, an `instrucciones` sheet in Spanish, and Excel data-validation dropdowns
for enum columns. The header order and names are derived directly from the typed
models so the template can never drift from the schema.
"""

from __future__ import annotations

import types
import typing
from enum import EnumMeta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from nexo_os.data.schema.models import (
    MODEL_BY_TABLE,
    OPERATIONAL_TABLES,
    OPTIONAL_TABLES,
    PII_FIELDS,
    _Row,
)

TEMPLATE_NAME = "nexo_carga_operativa.xlsx"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_OPTIONAL_FILL = PatternFill("solid", fgColor="BDD7EE")


def _enum_class(annotation: object) -> EnumMeta | None:
    """Return the StrEnum class for a field annotation (handling `Enum | None`)."""
    candidates = [annotation]
    if isinstance(annotation, types.UnionType) or typing.get_origin(annotation) is typing.Union:
        candidates = list(typing.get_args(annotation))
    for c in candidates:
        if isinstance(c, EnumMeta):
            return c
    return None


def _columns(model_cls: type[_Row]) -> list[str]:
    return list(model_cls.model_fields.keys())


def build_template(path: Path) -> Path:
    """Write the blank canonical workbook to `path`."""
    path = Path(path)
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet

    _add_instrucciones(wb)

    for table in OPERATIONAL_TABLES:
        model_cls = MODEL_BY_TABLE[table]
        ws = wb.create_sheet(title=table)
        cols = _columns(model_cls)
        for idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=idx, value=col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            ws.column_dimensions[get_column_letter(idx)].width = max(14, len(col) + 2)

            enum_cls = _enum_class(model_cls.model_fields[col].annotation)
            if enum_cls is not None:
                values = ",".join(m.value for m in enum_cls)
                if len(values) <= 250:  # Excel inline-list limit
                    dv = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
                    ws.add_data_validation(dv)
                    letter = get_column_letter(idx)
                    dv.add(f"{letter}2:{letter}1048576")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _add_instrucciones(wb: Workbook) -> None:
    ws = wb.create_sheet(title="instrucciones")
    ws.column_dimensions["A"].width = 110
    lines = [
        ("Nexo - Plantilla de carga operativa", True),
        ("", False),
        (
            "Complete una fila por registro en cada hoja. Los encabezados (fila 1) "
            "NO se deben modificar ni reordenar.",
            False,
        ),
        (
            "La carga es 'todo o nada': si el archivo tiene cualquier error de "
            "validacion, se rechaza completo y NO se ingiere nada. El snapshot "
            "anterior queda intacto. Corrija y vuelva a subir.",
            False,
        ),
        ("", False),
        ("Hojas requeridas:", True),
    ]
    for line, bold in lines:
        ws.append([line])
        if bold:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    for table in OPERATIONAL_TABLES:
        optional = table in OPTIONAL_TABLES
        tag = " (OPCIONAL - puede omitirse)" if optional else ""
        pii = PII_FIELDS.get(table)
        pii_note = f"  | Campos con datos personales (PII): {', '.join(sorted(pii))}" if pii else ""
        ws.append([f"  - {table}{tag}{pii_note}"])
        if optional:
            ws.cell(row=ws.max_row, column=1).fill = _OPTIONAL_FILL

    ws.append([""])
    ws.append(["Reglas clave:"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for rule in [
        "  - Montos en ARS, sin separador de miles. Use punto decimal. No se permiten montos negativos.",
        "  - Fechas en formato AAAA-MM-DD. fecha_fin_vigencia debe ser >= fecha_inicio_vigencia.",
        "  - Los IDs referenciados (cliente_id, aseguradora_id, productor_id, poliza_id, lead_id) deben existir.",
        "  - Las columnas con lista desplegable solo aceptan los valores permitidos.",
        "  - dias_mora, bucket_mora y diferencia_ars NO se cargan: Nexo los calcula contra la fecha del snapshot.",
    ]:
        ws.append([rule])
