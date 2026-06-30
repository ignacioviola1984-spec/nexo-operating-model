"""Phase 2: the canonical template matches the schema and has enum dropdowns."""

from __future__ import annotations

from openpyxl import load_workbook

from nexo_os.data.schema.models import MODEL_BY_TABLE, OPERATIONAL_TABLES
from nexo_os.data.template import build_template


def test_template_sheets_and_headers_match_schema(tmp_path):
    path = build_template(tmp_path / "tpl.xlsx")
    wb = load_workbook(path)
    assert set(wb.sheetnames) == set(OPERATIONAL_TABLES) | {"instrucciones"}

    for table in OPERATIONAL_TABLES:
        ws = wb[table]
        headers = [c.value for c in ws[1]]
        assert headers == list(MODEL_BY_TABLE[table].model_fields.keys())


def test_template_has_enum_dropdown(tmp_path):
    path = build_template(tmp_path / "tpl.xlsx")
    wb = load_workbook(path)
    # polizas.ramo is an enum column -> must carry a list data-validation.
    ws = wb["polizas"]
    assert len(ws.data_validations.dataValidation) >= 1
    kinds = {dv.type for dv in ws.data_validations.dataValidation}
    assert "list" in kinds
